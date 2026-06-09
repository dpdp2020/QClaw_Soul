# -*- coding: utf-8 -*-
"""
更新Excel图片URL并同步显示图片
用法: python update_excel_pics.py <excel路径> <图片列> <新URL或本地路径>
示例: python update_excel_pics.py test.xlsx H:J https://cbu01.alicdn.com/img/xxx.jpg
      python update_excel_pics.py test.xlsx H,J,K,L,M https://cbu01.alicdn.com/img/xxx.jpg
      python update_excel_pics.py test.xlsx aw168_img_16.jpg:本地路径 aw168_img_17.jpg:本地路径
"""
import sys, os, json, openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import urllib.request, hashlib

def parse_cols(col_spec):
    """解析列范围，如 H:J 或 H,J,K"""
    if ':' in col_spec:
        start, end = col_spec.split(':')
        start_idx = openpyxl.utils.column_index_from_string(start.upper())
        end_idx = openpyxl.utils.column_index_from_string(end.upper())
        return list(range(start_idx, end_idx + 1))
    elif ',' in col_spec:
        return [openpyxl.utils.column_index_from_string(c.strip().upper()) for c in col_spec.split(',')]
    else:
        return [openpyxl.utils.column_index_from_string(col_spec.upper())]

def download_img(url_or_path, tmpdir):
    """下载或复制图片到临时目录，返回本地路径"""
    if not url_or_path:
        return None
    if os.path.exists(url_or_path):
        # 本地文件，直接返回
        return url_or_path
    
    # URL下载
    if not str(url_or_path).startswith('http'):
        return None
    
    key = hashlib.md5(str(url_or_path).encode()).hexdigest()[:12]
    ext = '.jpg'
    if '.png' in str(url_or_path).lower():
        ext = '.png'
    fp = os.path.join(tmpdir, f'img_{key}{ext}')
    
    if not os.path.exists(fp):
        try:
            urllib.request.urlretrieve(str(url_or_path), fp)
        except:
            return None
    return fp

def update_pics(excel_path, col_spec, url_or_paths, row_start=2):
    """更新指定列的图片"""
    TMPDIR = os.path.join(os.environ.get('TEMP', 'C:\\temp'), '_xl_pics')
    os.makedirs(TMPDIR, exist_ok=True)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    col_indices = parse_cols(col_spec)
    
    # url_or_paths 可以是单个URL（所有列相同）或多个URL（每列不同）
    if isinstance(url_or_paths, str):
        urls = [url_or_paths] * len(col_indices)
    else:
        urls = url_or_paths
    
    # 先清除旧的图片对象（openpyxl会叠加）
    # 注：openpyxl无法直接删除图片，只能重新保存
    
    for row in ws.iter_rows(min_row=row_start):
        if not row[0].value:
            continue
        
        for col_idx, url in zip(col_indices, urls):
            cell = ws.cell(row=row[0].row, column=col_idx)
            cell.value = url
            
            # 下载并嵌入图片
            local = download_img(url, TMPDIR)
            if local and os.path.exists(local):
                ws.row_dimensions[cell.row].height = 110
                img = XLImage(local)
                img.width = 150
                img.height = 150
                img.anchor = cell.coordinate
                ws.add_image(img)
                print(f'  Row {cell.row} Col {get_column_letter(col_idx)}: OK')
            else:
                print(f'  Row {cell.row} Col {get_column_letter(col_idx)}: URL updated, image not embedded')
    
    # 保存
    wb.save(excel_path)
    print(f'\nExcel updated: {excel_path}')
    return excel_path

def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    
    excel_path = sys.argv[1]
    col_spec = sys.argv[2]
    
    # 剩余参数是URL或本地路径
    url_or_paths = sys.argv[3:]
    
    update_pics(excel_path, col_spec, url_or_paths)

if __name__ == '__main__':
    main()
